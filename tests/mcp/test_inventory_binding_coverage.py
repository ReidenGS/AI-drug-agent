"""Audit LocalMCP binding coverage against ToolUniversity_inventory_v0.2.xlsx."""

from __future__ import annotations

import os
from pathlib import Path

import pytest

from app.mcp.client import LocalMCPClient
from app.mcp.tools._registry import _all_bindings
from app.services.tool_inventory_service import ToolInventoryService


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = PROJECT_ROOT.parent / "ToolUniversity_inventory_v0.2.xlsx"


@pytest.fixture
def inventory() -> ToolInventoryService:
    xlsx = Path(os.environ.get("TOOL_INVENTORY_XLSX", str(DEFAULT_XLSX)))
    if not xlsx.exists():
        pytest.skip(f"Inventory xlsx not available at {xlsx}")
    return ToolInventoryService(str(xlsx))


def test_inventory_callable_tools_are_fully_bound(inventory):
    inventory_names = _callable_inventory_names(inventory)
    bound_names = {name for name, _ in _all_bindings()}
    missing = inventory_names - bound_names
    extras = bound_names - inventory.names()
    assert len(inventory_names) == 99
    assert len(bound_names) == 99
    assert not missing
    assert not extras


def test_local_mcp_scope_uses_inventory_filtered_registry(inventory):
    client = LocalMCPClient(inventory=inventory)
    all_bound = {name for name, _ in _all_bindings()}
    step5_tools = set(client.list_tools(agent_name="candidate_context_agent", step_id="step_05"))
    step13_tools = set(client.list_tools(agent_name="evidence_agent", step_id="step_13"))
    assert step5_tools
    assert step13_tools
    assert step5_tools <= all_bound
    assert step13_tools <= all_bound
    assert "MultiAgentLiteratureSearch" not in step5_tools
    assert "SAbDab_search_structures" not in step13_tools


def _callable_inventory_names(inventory: ToolInventoryService) -> set[str]:
    from openpyxl import load_workbook

    wb = load_workbook(inventory.xlsx_path, data_only=True)
    ws = wb["ToolUniversity_inventory"]
    header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    out: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[idx["Tool Name"]]
        in_tu = row[idx["Is In ToolUniverse"]]
        if str(in_tu or "").strip().lower() != "yes":
            continue
        if not name or str(name).strip().lower() == "n/a":
            continue
        out.add(str(name).strip())
    return out
