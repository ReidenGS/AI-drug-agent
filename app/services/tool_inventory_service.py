"""Import ToolUniversity_inventory_v0.2.xlsx into the tool registry.

This is the SINGLE source of MCP-callable tools per the architecture document
(`README_FOR_CLAUDE.md` hard constraint: "Do not register full ToolUniverse
all-tools extract"). FastMCP wires only tools whose name appears here.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


@dataclass(slots=True)
class InventoryEntry:
    tool_name: str
    step_id: str | None
    pipeline_stage: str | None
    tool_status: str | None      # buildable | external | wrapper | available | future
    runtime_status: str | None
    category: str | None
    notes: str | None


class ToolInventoryService:
    def __init__(self, xlsx_path: str) -> None:
        self.xlsx_path = Path(xlsx_path)

    def load(self) -> list[InventoryEntry]:
        from openpyxl import load_workbook

        wb = load_workbook(self.xlsx_path, data_only=True)
        ws = wb.active
        header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(header)}

        def col(row: tuple, name: str) -> str | None:
            i = idx.get(name)
            if i is None:
                return None
            v = row[i]
            return None if v is None else str(v).strip()

        entries: list[InventoryEntry] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            tool_name = col(row, "tool_name") or col(row, "Tool Name")
            if not tool_name:
                continue
            entries.append(
                InventoryEntry(
                    tool_name=tool_name,
                    step_id=col(row, "step_id") or col(row, "Step"),
                    pipeline_stage=col(row, "pipeline_stage") or col(row, "Pipeline Stage"),
                    tool_status=col(row, "tool_status") or col(row, "Status"),
                    runtime_status=col(row, "runtime_status"),
                    category=col(row, "category") or col(row, "Category"),
                    notes=col(row, "notes") or col(row, "Notes"),
                )
            )
        return entries

    def names(self) -> set[str]:
        return {e.tool_name for e in self.load()}

    def scope_for(
        self,
        *,
        agent_name: str | None = None,
        step_id: str | None = None,
        statuses: Iterable[str] | None = None,
    ) -> list[InventoryEntry]:
        items = self.load()
        if step_id:
            items = [e for e in items if (e.step_id or "").startswith(step_id)]
        if statuses:
            allow = set(statuses)
            items = [e for e in items if (e.tool_status or "") in allow]
        return items
